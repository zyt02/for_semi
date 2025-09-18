#analysis.py

import pandas as pd
import os
from io import BytesIO
from openpyxl.styles import Font
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl import load_workbook


#output_path = '/Users/mac/Downloads/highlighted_rows.xlsx'

def analyze_taxi_expense(taxi_df):
    taxi_df["date"] = pd.to_datetime(taxi_df["交易時間"]).dt.date
    summary = taxi_df.groupby(["姓名", "date"])["乘車券編號"].count().reset_index()
    result = pd.merge(taxi_df, summary, how = 'left', on = ["姓名", "date"])
    result = result.rename(columns={"乘車券編號_y" : "當日搭乘次數"}) 
    result = result.sort_values(by = ["姓名", "交易時間"]).reset_index()

    def highlight_rows(row):
        return ['background-color:yellow' if row['當日搭乘次數'] >= 3 else '' for _ in row ]
    
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if result.empty: 
            print("no result")
        else: 
            result.style.apply(highlight_rows, axis=1).to_excel(writer, index=False, sheet_name='Sheet1')
            print(f"File saved to: {output}")
    output.seek(0)
    return output.getvalue()


#output_path = '/Users/mac/Downloads/late_early_rows.xlsx'

def analyze_schedule_status(log_df):
    # data prep
    # save as excel workbook (xlsx) not Strict Open XML Spreadsheet (xlsx)
    log_df['刷卡日期'] = pd.to_datetime(log_df['刷卡日期']).dt.date
    log_df['最早刷卡時間'] = pd.to_datetime(log_df['最早刷卡時間'], format='%H:%M:%S').dt.time
    log_df['最晚刷卡時間'] = pd.to_datetime(log_df['最晚刷卡時間'], format='%H:%M:%S').dt.time
    log_df['shift_start_30min'] = pd.to_datetime(log_df['shift_start_30min'], format='%H:%M:%S').dt.time
    log_df['shift_end_30min'] = pd.to_datetime(log_df['shift_end_30min'], format='%H:%M:%S').dt.time

    log_details = log_df

    # Define fonts (you need to do this inside the function)
    orange_font = Font(color="F08000") 
    yellow_font = Font(color="FFC000") 
    red_font = Font(color= "C70039")

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if log_details.empty:
            print("No results to write.")
            return None  # Or handle however you like
        else:
            log_details.to_excel(writer, index=False, sheet_name='Sheet1')

            workbook = writer.book
            sheet = workbook['Sheet1']

            header_row = [cell.value for cell in sheet[1]]  # First row headers

            try:
                col_idx_earliest_swipe = header_row.index("最早刷卡時間") + 1
                col_idx_latest_swipe = header_row.index("最晚刷卡時間") + 1
                col_idx_name = header_row.index("姓名") + 1
                col_idx_date = header_row.index("刷卡日期") + 1
            except ValueError as e:
                print(f"Error: One of the required columns not found in the DataFrame. {e}")
                return None

            for i, row in log_details.iterrows():
                excel_row_num = i + 2  # Excel rows start at 1, plus header row

                # Style logic
                if row["最早刷卡時間"] >= row['shift_start_30min']:
                    sheet.cell(row=excel_row_num, column=col_idx_earliest_swipe).font = orange_font
                    sheet.cell(row=excel_row_num, column=col_idx_name).font = red_font
                    sheet.cell(row=excel_row_num, column=col_idx_date).font = red_font

                if row["最晚刷卡時間"] <= row['shift_end_30min']:
                    sheet.cell(row=excel_row_num, column=col_idx_latest_swipe).font = yellow_font
                    sheet.cell(row=excel_row_num, column=col_idx_name).font = red_font
                    sheet.cell(row=excel_row_num, column=col_idx_date).font = red_font

    output.seek(0)
    return output.getvalue()
